#!/usr/bin/env -S uv run --quiet --no-upgrade --script
# /// script
# requires-python = "~=3.12"
# dependencies = [
# "click",
# "xlwings",
# "python-decouple",
# "pandas",
# "pypdf",
# "docx2pdf",
# "reportlab",
# "openpyxl",
# "playwright",
# ]
# ///

# For the import, remember to put it in the script above. Script is read during runtime.
# Heavy imports (docx2pdf, xlwings, pypdf, etc.) are lazy-loaded in commands that need them
import logging
import math
import os
import tomllib
import platform
import re
import shutil
import subprocess
import getpass
from datetime import datetime
from pathlib import Path

import click

# Version
# Recommended is to define in __init__.py but I am doing it here.
__version__ = "0.1.0"

# Handle case for different users
username = getpass.getuser()
match username:
    case "oliver":
        RFQ = "~/OneDrive - Jason Electronics Pte Ltd/Shared Documents/@rfqs/"
        HO = "~/OneDrive - Jason Electronics Pte Ltd/Shared Documents/@handover/"
        CO = "~/OneDrive - Jason Electronics Pte Ltd/Shared Documents/@costing/"
        DOCS = "~/OneDrive - Jason Electronics Pte Ltd/Shared Documents/@docs/"
        TOOLS = "~/OneDrive - Jason Electronics Pte Ltd/Shared Documents/@tools/"
        BID_ALIAS = f"alias bid=\"uv run --quiet '{Path(r'~/OneDrive - Jason Electronics Pte Ltd/Shared Documents/@tools/bid.py').expanduser().resolve()}'\""

    case "carol_lim":
        RFQ = "~/Jason Electronics Pte Ltd/Bid Proposal - Documents/@rfqs/"
        DOCS = "~/Jason Electronics Pte Ltd/Bid Proposal - Documents/@docs/"
        TOOLS = "~/Jason Electronics Pte Ltd/Bid Proposal - @tools/"
        BID_ALIAS = f"alias bid=\"uv run --quiet '{Path(r'~/Jason Electronics Pte Ltd/Bid Proposal - @tools/bid.py').expanduser().resolve()}'\""

    case _:
        RFQ = "~/Jason Electronics Pte Ltd/Bid Proposal - Documents/@rfqs/"
        HO = "~/Jason Electronics Pte Ltd/Bid Proposal - Documents/@handover/"
        CO = "~/Jason Electronics Pte Ltd/Bid Proposal - Documents/@costing/"
        DOCS = "~/Jason Electronics Pte Ltd/Bid Proposal - Documents/@docs/"
        TOOLS = "~/Jason Electronics Pte Ltd/Bid Proposal - Documents/@tools/"
        BID_ALIAS = f"alias bid=\"uv run --quiet '{Path(r'~/Jason Electronics Pte Ltd/Bid Proposal - Documents/@tools/bid.py').expanduser().resolve()}'\""


RESTRICTED_FOLDER = [
    "@rfqs",
    "@costing",
    "@handover",
    "@tools",
    "@commercial-review",
    "@projects",
    "Documents",
]

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)


@click.command()
@click.option("-y", "--yes", is_flag=True, help="Answer yes to the current directory")
def word2pdf(yes: bool):
    """
    (Batch) Convert word file to pdf in given directory.
    Require MS Word to be installed as it is doing the conversion.
    """
    import docx2pdf

    directory = Path.cwd()
    if not yes:
        click.confirm(
            f"The command will convert all the word files (only .docx (case sensitive)) in '{directory}'",
            abort=True,
        )
    docx2pdf.convert(directory)


def normalize_job_code(search_term: str) -> str:
    """Normalize job code input.

    If only 3 digits are given, prepend 'J12' to make a full job code.
    E.g., '786' becomes 'J12786'
    """
    search_term = search_term.strip()
    if search_term.isdigit() and len(search_term) == 3:
        return f"J12{search_term}"
    return search_term


def find_project_folder(search_term: str) -> list[tuple[Path, int]]:
    """Search for project folders matching the search term.

    Searches in @rfqs/<year>/ for years from current year down to 2023.
    Returns list of (folder_path, year) tuples for folders starting with search_term.
    """
    search_term = normalize_job_code(search_term)
    rfqs_path = Path(RFQ).expanduser()

    if not rfqs_path.exists():
        return []

    matches = []
    current_year = datetime.now().year

    for year in range(current_year, 2022, -1):  # Down to 2023
        year_path = rfqs_path / str(year)
        if not year_path.exists():
            continue

        for folder in year_path.iterdir():
            if folder.is_dir() and folder.name.upper().startswith(search_term.upper()):
                matches.append((folder, year))

    return matches


def get_next_vo_number(vo_folder: Path) -> int:
    """Get the next available VO number.

    Scans existing NN-VO folders and returns the next available number.
    """
    if not vo_folder.exists():
        return 1

    existing_numbers = []
    for folder in vo_folder.iterdir():
        if folder.is_dir():
            # Match pattern like "01-VO", "02-VO", etc.
            match = re.match(r"^(\d{2})-VO\b", folder.name)
            if match:
                existing_numbers.append(int(match.group(1)))

    if not existing_numbers:
        return 1

    return max(existing_numbers) + 1


def create_vo_structure(vo_path: Path) -> None:
    """Create the VO subfolder structure."""
    iso_date = datetime.now().strftime("%Y-%m-%d")

    folders = [
        vo_path / "00-ITB" / iso_date,
        vo_path / "01-Commercial" / "00-Arc",
        vo_path / "02-Technical" / "00-Arc",
        vo_path / "03-Supplier" / "00-Arc",
        vo_path / "04-Datasheet" / "00-Arc",
        vo_path / "05-PO" / "00-Arc",
    ]

    for folder in folders:
        folder.mkdir(parents=True, exist_ok=True)


def get_handover_candidates(project_path: Path) -> list[tuple[str, Path]]:
    """Get list of handover candidates for a project.

    Returns list of (display_name, source_path) tuples:
    - Main project folder as "00-MAIN"
    - Any NN-VO folders from 07-VO/
    """
    candidates = [("00-MAIN", project_path)]

    vo_parent = project_path / "07-VO"
    if vo_parent.exists():
        vo_folders = sorted(
            [
                f
                for f in vo_parent.iterdir()
                if f.is_dir() and re.match(r"^\d{2}-VO\b", f.name)
            ],
            key=lambda x: x.name,
        )
        for vo_folder in vo_folders:
            candidates.append((vo_folder.name, vo_folder))

    return candidates


def is_folder_not_empty(path: Path) -> bool:
    """Check if folder exists and is not empty."""
    if not path.exists():
        return False
    return any(path.iterdir())


def is_conforming_handover_folder(folder: Path) -> bool:
    """Check if folder was created by bid ho (has expected structure).

    Returns True if folder doesn't exist (new folder will conform) or
    if it has at least 3 of the expected subfolders.
    """
    if not folder.exists():
        return True  # New folder, will conform

    expected = {
        "00-ITB",
        "01-PO",
        "02-Technical",
        "03-Supplier",
        "04-Datasheet",
        "05-Cost",
    }
    existing = {f.name for f in folder.iterdir() if f.is_dir()}

    # Check if it has at least 3 expected folders
    return len(expected & existing) >= 3


def _copy_with_retry(source: Path, dest: Path, max_retries: int = 3) -> None:
    """Copy a file with retry on timeout.

    OneDrive Files On-Demand (cold files) need to be hydrated from SharePoint
    on first access, which can cause TimeoutError on slow or cold connections.
    Retries with exponential backoff to give OneDrive time to connect.
    """
    import time

    for attempt in range(max_retries):
        try:
            shutil.copy2(source, dest)
            return
        except (TimeoutError, OSError) as e:
            if attempt < max_retries - 1:
                wait = 15 * (attempt + 1)
                click.echo(
                    f"  Timeout copying {source.name}, retrying in {wait}s "
                    f"(attempt {attempt + 2}/{max_retries})..."
                )
                time.sleep(wait)
            else:
                raise


def sync_folder(source: Path, dest: Path) -> bool:
    """One-way mirror sync from source to dest.

    Returns True if sync was performed, False if source was empty/missing.
    Mirrors source to dest: adds new files, updates changed files, deletes
    files in dest that don't exist in source. Empty folders are skipped.
    """
    if not is_folder_not_empty(source):
        return False

    dest.mkdir(parents=True, exist_ok=True)

    # Get only files (not directories) - empty folders like 00-Arc are skipped
    source_files = {p.relative_to(source) for p in source.rglob("*") if p.is_file()}
    dest_files = {p.relative_to(dest) for p in dest.rglob("*") if p.is_file()}

    # Delete files in dest that don't exist in source (purge)
    for rel_path in dest_files - source_files:
        dest_path = dest / rel_path
        if dest_path.exists():
            dest_path.unlink()

    # Clean up empty directories in dest
    for dir_path in sorted(dest.rglob("*"), reverse=True):
        if dir_path.is_dir() and not any(dir_path.iterdir()):
            dir_path.rmdir()

    # Copy/update files from source to dest
    for rel_path in source_files:
        source_path = source / rel_path
        dest_path = dest / rel_path

        # Copy if dest doesn't exist or source is newer
        if (
            not dest_path.exists()
            or source_path.stat().st_mtime > dest_path.stat().st_mtime
        ):
            dest_path.parent.mkdir(parents=True, exist_ok=True)
            _copy_with_retry(source_path, dest_path)

    return True


def perform_handover_sync(source_root: Path, dest_root: Path, is_vo: bool) -> None:
    """Perform the handover sync operations.

    Args:
        source_root: Source folder (project root or VO folder)
        dest_root: Destination folder in @handover
        is_vo: True if this is a VO handover (affects PO folder mapping)
    """
    dest_root.mkdir(parents=True, exist_ok=True)

    # Define sync mappings: (source_name, dest_name)
    # PO folder differs: main project has 06-PO, VO has 05-PO
    po_source = "05-PO" if is_vo else "06-PO"

    sync_mappings = [
        ("00-ITB", "00-ITB"),
        (po_source, "01-PO"),
        ("02-Technical", "02-Technical"),
        ("03-Supplier", "03-Supplier"),
        ("04-Datasheet", "04-Datasheet"),
        ("05-Drawing", "06-Drawing"),
    ]

    for source_name, dest_name in sync_mappings:
        source_path = source_root / source_name
        dest_path = dest_root / dest_name

        if sync_folder(source_path, dest_path):
            click.echo(f"  Synced: {source_name} -> {dest_name}")
        else:
            click.echo(f"  Skipped: {source_name} (empty or missing)")

    # Always create 05-Cost folder
    cost_folder = dest_root / "05-Cost"
    cost_folder.mkdir(parents=True, exist_ok=True)
    click.echo("  Created: 05-Cost")


def open_with_default_app(file_path: Path):
    "Open file_path with default application"
    if platform.system() == "Windows":
        try:
            subprocess.run(["explorer", file_path.expanduser().resolve()])
        except Exception as _:
            click.echo("Not yet implemented.")
    elif platform.system() == "Darwin":
        click.launch("file://" + str(file_path.expanduser().resolve()))
    else:
        click.echo("Not implemented for this platform")


def remove_folder(folder_path):
    if os.path.exists(folder_path):
        try:
            shutil.rmtree(folder_path)
            click.echo(f"Deleted: {folder_path}")
            return True
        except OSError as e:
            click.echo(f"Error deleting folder: {e}")
            return False
    else:
        click.echo(f"Folder {folder_path} does not exist.")


def require_rename(file_name: str, flag: bool = False) -> tuple[str, bool]:
    """
    Check if file meets required naming specs
    If not, suggest renaming and the name

    Specs:
    Strip file name of white space at the end
    Put file extension in lowercase
    `word-` to `word -`
    One or more '-' to single '-'
    One or more '_' to single '_'
    One or more ` . ` to single ` `
    ` _word` to ` word`
    One or more `#`, `_` followed by space to single space
    Delete `[` or `]`
    Remove RE, SV, FW, FWD, EXTERNAL, URGENT, 回复 at the start of email, case insensitive
    No double 'space' or more

    """

    file, extension = os.path.splitext(file_name)
    new_file_name = file.strip() + extension.lower()

    new_file_name = re.sub(r"(\b\w+\b)-\s", r"\1 - ", new_file_name)
    # new_file_name = re.sub(r"-{1,}", " ", new_file_name)
    new_file_name = re.sub(r"-{2,}", "-", new_file_name)
    new_file_name = re.sub(r"_{2,}", "_", new_file_name)
    new_file_name = re.sub(r" \.+ ", " ", new_file_name)
    new_file_name = re.sub(r" _(\w+)", r" \1", new_file_name)
    new_file_name = re.sub(r"(#+|_+\s{1,})", r" ", new_file_name)
    new_file_name = re.sub(r"(\[+)", "", new_file_name)
    new_file_name = re.sub(r"(\]+)", "", new_file_name)
    new_file_name = re.sub(r"^(RE(_+|\s{1,}))", "", new_file_name, flags=re.IGNORECASE)
    new_file_name = re.sub(r"^(SV(_+|\s{1,}))", "", new_file_name, flags=re.IGNORECASE)
    new_file_name = re.sub(r"^(AW(_+|\s{1,}))", "", new_file_name, flags=re.IGNORECASE)
    new_file_name = re.sub(r"^(FW(_+|\s{1,}))", "", new_file_name, flags=re.IGNORECASE)
    new_file_name = re.sub(r"^(FWD(_+|\s{1,}))", "", new_file_name, flags=re.IGNORECASE)
    new_file_name = re.sub(
        r"^(回复(_+|\s{1,}))", "", new_file_name, flags=re.IGNORECASE
    )
    new_file_name = re.sub(
        r"^(URGENT(_+|\s{1,}))", "", new_file_name, flags=re.IGNORECASE
    )
    new_file_name = re.sub(
        r"^(EXTERNAL(_+|\s{1,}))", "", new_file_name, flags=re.IGNORECASE
    )
    new_file_name = re.sub(
        r"^(EXTERNALRE(_+|\s{1,}))", "", new_file_name, flags=re.IGNORECASE
    )
    new_file_name = re.sub(r"^(FWD(_+|\s{1,}))", "", new_file_name, flags=re.IGNORECASE)
    new_file_name = re.sub(r"^(FW(_+|\s{1,}))", "", new_file_name, flags=re.IGNORECASE)
    new_file_name = re.sub(r"^(SV(_+|\s{1,}))", "", new_file_name, flags=re.IGNORECASE)
    new_file_name = re.sub(r"^(RE(_+|\s{1,}))", "", new_file_name, flags=re.IGNORECASE)
    # new_file_name = re.sub(
    #     r"(-+|\s{1,})", " ", new_file_name, flags=re.IGNORECASE
    # )  # remove hyphen
    new_file_name = re.sub(r"\s{2,}", " ", new_file_name)
    if new_file_name != file_name:
        flag = True
        return (new_file_name, flag)
    return (file_name, flag)


def rename_file(old_file_name: Path, new_file_name: Path) -> bool:
    "Rename old_file_name to new_file_name"
    try:
        if os.path.exists(new_file_name):
            # Handle case where file with same extension but different case exists
            if new_file_name.suffix != old_file_name.suffix:
                os.rename(old_file_name, new_file_name)
                return True
            # Othereise skip
            click.echo(f"File with the same name as '{new_file_name}' already exists.")
            return False
        os.rename(old_file_name, new_file_name)
        return True
    except OSError as e:
        click.echo(f"Error renaming file: '{old_file_name}' caused by {e}")
        return False


# Create project folder structure
@click.command()
@click.argument("folder_name", default="")
def init(folder_name: str) -> None:
    """
    Create folder structure for project in @rfqs.
    Search for the latest year, create one if it does not exists.
    Then create the required folder structure in it.
    """
    from util import excelx

    # Handle case where @rfqs does not exists
    if not Path(RFQ).expanduser().exists():
        click.echo("The folder @rfqs does not exist. Check if you have access.")
        return

    if folder_name == "":
        folder_name = click.prompt("Please enter folder name to create")
    # Normalize dashes: convert en-dash (–) and em-dash (—) to hyphen-minus (-)
    folder_name = folder_name.replace("–", "-").replace("—", "-")
    current_year = datetime.now().year
    path = Path(RFQ).expanduser()
    folders = []
    for item in path.iterdir():
        if item.is_dir():
            folders.append(item.name)
    if str(current_year) in folders:
        new_path = path / Path(str(current_year))
    else:
        new_path = path / Path(str(current_year))
        new_path.mkdir(exist_ok=True)
    # Create subdirectories if the folder_name not exists
    new_path = new_path / folder_name
    if new_path.exists():
        click.echo("The folder already exists.")
        # Create new commercial Proposal
        if click.confirm(
            "Do you want to create a new Commercial Proposal?", default=True
        ):
            version = click.prompt(
                "Version No, default:", default="B0", show_default=True
            )
            version = str(version).upper()
            template_folder = (
                Path(RFQ).expanduser().parent.absolute().resolve() / "@tools/resources"
            )
            template = "Template.xlsx"
            commercial_folder = new_path / "01-Commercial"
            commercial_file = folder_name + " " + str(version) + ".xlsx"
            jobcode = folder_name.split()[0]
            excelx.create_excel_from_template(
                template_folder,
                template,
                commercial_folder,
                commercial_file,
                jobcode,
                version,
            )
        if click.confirm("Do you want to open the project folder?", abort=True):
            open_with_default_app(new_path)
            return
    else:
        if click.confirm(
            f"'{new_path}' will be created. Continue?", default=True, abort=True
        ):
            new_path.mkdir()
            iso_date = datetime.now().strftime("%Y-%m-%d")
            itb_specs = new_path / "00-ITB" / iso_date
            itb_specs.mkdir(parents=True, exist_ok=True)
            commercial = new_path / "01-Commercial/00-Arc"
            commercial.mkdir(parents=True, exist_ok=True)
            technical = new_path / "02-Technical/00-Arc"
            technical.mkdir(parents=True, exist_ok=True)
            supplier = new_path / "03-Supplier/00-Arc"
            supplier.mkdir(parents=True, exist_ok=True)
            datasheet = new_path / "04-Datasheet/00-Arc"
            datasheet.mkdir(parents=True, exist_ok=True)
            drawing = new_path / "05-Drawing/00-Arc"
            drawing.mkdir(parents=True, exist_ok=True)
            po = new_path / "06-PO/00-Arc"
            po.mkdir(parents=True, exist_ok=True)
            vo = new_path / "07-VO"
            vo.mkdir(parents=True, exist_ok=True)
            toolkit = new_path / "08-Toolkit"
            toolkit.mkdir(parents=True, exist_ok=True)
            click.echo(f"Created fodler {new_path}")
            # Create new commercial Proposal
            if click.confirm(
                "Do you want to create a new Commercial Proposal?", default=True
            ):
                version = click.prompt(
                    "Version No, default:", default="B0", show_default=True
                )
                version = str(version).upper()
                template_folder = (
                    Path(RFQ).expanduser().parent.absolute().resolve()
                    / "@tools/resources"
                )
                template = "Template.xlsx"
                commercial_folder = new_path / "01-Commercial"
                commercial_file = folder_name + " " + str(version) + ".xlsx"
                jobcode = folder_name.split()[0]
                excelx.create_excel_from_template(
                    template_folder,
                    template,
                    commercial_folder,
                    commercial_file,
                    jobcode,
                    version,
                )
            # Ask for the folder to be opened
            if click.confirm(
                "Do you want to open the project folder?",
                abort=True,
            ):
                open_with_default_app(new_path)
        else:
            click.echo("Folder creation aborted.")


@click.command()
@click.argument("folder_name", default="")
@click.option("-d", "--dry-run", is_flag=True, help="Dry run mode")
@click.option("-y", "--yes", is_flag=True, help="Answer yes to the current directory")
@click.option(
    "-r",
    "--remove-git",
    is_flag=True,
    help="Delete .git folder and .gitignore if exists",
)
def clean(folder_name: str, dry_run: bool, remove_git: bool, yes: bool) -> None:
    """
    Clean file names in folder. Default search path is @rfqs.
    If folder is given, clean the folder.
    If not, search the folder in @rfqs and clean.
    """
    if folder_name == "":
        if yes:
            folder_name = Path.cwd().name
        else:
            folder_name = click.prompt(
                (
                    "Please enter folder name to clean. "
                    "Default search path is in @rfqs. \n"
                    "Press `Enter` to choose current folder:"
                ),
                default=Path.cwd().name,
                type=str,
            )
    if (
        folder_name in RESTRICTED_FOLDER
        or folder_name.isdigit()  # Main folder in @rfqs
        or folder_name == Path.home().name
    ):
        click.echo(
            f"You are not allowed to clean '{folder_name}' directly. "
            "Please choose a subfolder."
        )
        return
    if folder_name == Path.cwd().name:
        clean_folder(Path.cwd(), dry_run=dry_run, yes=yes)
    else:
        clean_rfqs(folder_name, remove_git=remove_git, dry_run=dry_run)


def clean_folder(start_path, dry_run=False, yes=False):
    "Clean folder"
    rename_list = []
    for root, _, files in os.walk(start_path):
        for file in files:
            check = require_rename((file))
            if check[1]:
                rename_list.append((root, file, check[0]))
    if not dry_run:
        if not rename_list:
            click.echo("No file required to be renamed. Folder clean 😎")
            return
        click.echo("Here is the list of files to rename.")
        for item in rename_list:
            click.echo(f"{item[1]} -> {item[2]}")
        click.echo(f"{len(rename_list)} files to rename")
        if not yes:
            click.confirm("Do you want to rename the files?", abort=True)
        click.echo("Renaming files")
        count = 0
        for item in rename_list:
            check = rename_file(
                Path(item[0]).joinpath(item[1]), Path(item[0]).joinpath(item[2])
            )
            if check:
                count += 1
                click.echo(f"Renamed: '{item[1]}' -> '{item[2]}'")
            else:
                # Handle the case where file with same name exists
                # Append -000, try once
                new_file_name = Path(item[2]).stem + "-000" + Path(item[2]).suffix
                check_again = rename_file(
                    Path(item[0]).joinpath(item[1]),
                    Path(item[0]).joinpath(new_file_name),
                )
                if check_again:
                    count += 1
                    click.echo(
                        (
                            f"Since the same file name exists '{item[1]}' "
                            f"renamed to '{new_file_name}' instead appending '-000'"
                        )
                    )
                else:
                    click.echo(f"Failed to rename '{item[1]}'")
        click.echo(f"Total {count} files renamed.")
        return
    else:  # if dry_run
        if not rename_list:
            click.echo("No file required to be renamed. Folder clean 😎")
            return
        click.echo("Here is the list of files to rename.")
        for item in rename_list:
            click.echo(f"'{item[1]}' -> '{item[2]}'")
        click.echo(f"{len(rename_list)} files to rename")


def clean_rfqs(folder_name, remove_git=False, dry_run=False):
    "Search for folder name in @rfqs and clean it"
    # Handle case where @rfqs does not exists
    if not Path(RFQ).expanduser().exists():
        click.echo("The folder @rfqs does not exist. Check if you have access.")
        return
    rfqs = Path(RFQ).expanduser()
    folders = []
    for _, dirs, _ in os.walk(rfqs):
        for dir in dirs:
            folders.append(dir)
        dirs.clear()  # To stop at top level folder
        folders = [str(item) for item in folders if str(item).isdigit()]
        folders = sorted(folders, key=lambda x: int(x), reverse=True)

    # Walk files in the second level
    # rename_list = []
    for level in folders:
        for _, dirs, files in os.walk(rfqs / level):
            for dir in dirs:
                if folder_name == dir:
                    start_path = Path(rfqs / level / dir)
                    # Check for .git folder and .gitignore
                    for _, dirs, files in os.walk(start_path):
                        if ".git" in dirs:
                            git_path = Path(rfqs / level / dir / ".git")
                            click.echo(f"Found {git_path}")
                            if remove_git:
                                remove_folder(git_path)
                        if ".gitignore" in files:
                            gitignore_path = Path(rfqs / level / dir / ".gitignore")
                            click.echo(f"Found {gitignore_path}")
                            if remove_git:
                                try:
                                    os.remove(gitignore_path)
                                    click.echo(f"Deleted: {gitignore_path}")
                                except OSError as e:
                                    click.echo(f"Error deleting file: {e}")
                        dirs.clear()  # Stop at top level folder
                    # Now clean folder
                    clean_folder(start_path, dry_run=dry_run)
                    return
            dirs.clear()  # Stop at top level folder
    click.echo(f"{folder_name} cannot be found in {rfqs}")


@click.command()
def setup():
    """
    Setup necessary environment variables and alias.
    """
    if platform.system() == "Linux":
        pass
    elif platform.system() == "Windows":
        # Step 1 — Create .managed_python folder
        managed_python = Path.home() / ".managed_python"
        managed_python.mkdir(exist_ok=True)
        click.echo(f"Created {managed_python}")

        # Step 1b — Set UV_CACHE_DIR in .bashrc to avoid AV-locked AppData cache
        if username == "carol_lim":
            uv_cache = Path.home() / "uv-cache"
            uv_cache.mkdir(exist_ok=True)
            bashrc = Path.home() / ".bashrc"
            export_line = 'export UV_CACHE_DIR="$HOME/uv-cache"'
            existing = bashrc.read_text() if bashrc.exists() else ""
            if "UV_CACHE_DIR" not in existing:
                with open(bashrc, "a") as f:
                    f.write(f"\n{export_line}\n")
                click.echo(f"Added UV_CACHE_DIR to {bashrc}")
            else:
                click.echo("UV_CACHE_DIR already set in .bashrc, skipping.")

        # Step 2 — Copy pyproject.toml from @tools (use script's own directory)
        tools_path = Path(__file__).parent
        shutil.copy2(tools_path / "pyproject.toml", managed_python / "pyproject.toml")
        click.echo("Copied pyproject.toml to .managed_python.")

        # Step 3 — Run uv sync (skip if venv already exists)
        venv_python = managed_python / ".venv" / "Scripts" / "python.exe"
        if venv_python.exists():
            click.echo("Virtual environment already exists, skipping uv sync.")
        else:
            click.echo("Running uv sync...")
            subprocess.run(["uv", "sync"], cwd=str(managed_python), check=True)
            click.echo("uv sync complete.")

        # Step 4 — Add .venv/Scripts Python to user PATH (top priority)
        python_path = str(managed_python / ".venv" / "Scripts")
        result = subprocess.run(
            [
                "powershell",
                "-c",
                "[Environment]::GetEnvironmentVariable('Path', 'User')",
            ],
            capture_output=True,
            text=True,
        )
        current_path = result.stdout.strip()
        if python_path.lower() not in current_path.lower():
            new_path = python_path + ";" + current_path
            subprocess.run(
                [
                    "powershell",
                    "-c",
                    f"[Environment]::SetEnvironmentVariable('Path', '{new_path}', 'User')",
                ],
                check=True,
            )
            click.echo(f"Added {python_path} to user PATH.")
        else:
            click.echo(f"{python_path} is already in user PATH, skipping.")

        # Check if Excel is running before steps 5 & 6 (both require Excel closed)
        excel_closed = True
        result = subprocess.run(
            ["powershell", "-c", "Get-Process excel -ErrorAction SilentlyContinue"],
            capture_output=True,
            text=True,
        )
        if result.returncode == 0:
            if click.confirm("Excel is running. Close it to continue setup?"):
                subprocess.run(["powershell", "-c", "Stop-Process -Name excel -Force"])
            else:
                click.echo("Skipping xlwings add-in install and PERSONAL.XLSB copy.")
                excel_closed = False

        if excel_closed:
            # Step 5 — Install xlwings add-in
            click.echo("Installing xlwings add-in...")
            subprocess.run(
                ["xlwings", "addin", "install"],
                cwd=str(managed_python),
                check=True,
            )
            click.echo("xlwings add-in installed.")

            # Step 6 — Copy PERSONAL.XLSB to XLSTART
            personal_xlsb_src = tools_path / "PERSONAL.XLSB"
            xlstart = (
                Path.home() / "AppData" / "Roaming" / "Microsoft" / "Excel" / "XLSTART"
            )
            personal_xlsb_dst = xlstart / "PERSONAL.XLSB"
            xlstart.mkdir(parents=True, exist_ok=True)
            shutil.copy2(personal_xlsb_src, personal_xlsb_dst)
            click.echo(f"Copied PERSONAL.XLSB to {xlstart}")

        # Step 7 — Ensure PERSONAL.XLSB opens hidden when Excel starts
        # This is controlled by VBA in PERSONAL.XLSB's ThisWorkbook module:
        #   Private Sub Workbook_Open()
        #       Windows(ThisWorkbook.Name).Visible = False
        #   End Sub
        # No action needed here — the file in @tools is pre-configured.

        # Step 8 — Make .managed_python folder hidden in Windows
        subprocess.run(
            ["attrib", "+H", str(managed_python)],
            check=True,
        )
        click.echo(f"Set {managed_python} as hidden.")

        # Step 9 — Set xlwings interpreter path to .managed_python/.venv python
        xlwings_conf_dir = Path.home() / ".xlwings"
        xlwings_conf_dir.mkdir(exist_ok=True)
        xlwings_conf = xlwings_conf_dir / "xlwings.conf"
        interpreter_path = str(managed_python / ".venv" / "Scripts" / "python.exe")
        pythonpath = str(tools_path).replace("@", "\\@")

        # Step 10 — Set xlwings PYTHONPATH to @tools folder and disable ADD_WORKBOOK_TO_PYTHONPATH
        # xlwings.conf uses CSV format: "KEY","VALUE"
        # Must use newline="" on both read and write to prevent \r\n → \r\r\n
        # double conversion, which causes VBA "Input past end of file" error 62
        import csv

        conf_lines = {}
        if xlwings_conf.exists():
            with open(xlwings_conf, newline="") as f:
                for row in csv.reader(f):
                    if len(row) >= 2:
                        conf_lines[row[0]] = row[1]
        conf_lines["INTERPRETER_WIN"] = interpreter_path
        conf_lines["PYTHONPATH"] = pythonpath
        conf_lines["ADD_WORKBOOK_TO_PYTHONPATH"] = "False"
        with open(xlwings_conf, "w", newline="") as f:
            writer = csv.writer(f, quoting=csv.QUOTE_ALL)
            for k, v in conf_lines.items():
                writer.writerow([k, v])
        click.echo(f"Set xlwings interpreter to {interpreter_path}")
        click.echo(f"Set xlwings PYTHONPATH to {pythonpath}")
        click.echo("Disabled ADD_WORKBOOK_TO_PYTHONPATH.")

        # Step 11 — Copy Excel.officeUI ribbon customization
        # Replace hardcoded XLSTART path in onAction attributes with current user's
        # path (onAction needs the full path for application-level ribbon callbacks)
        office_ui_src = tools_path / "resources" / "Excel.officeUI"
        office_ui_dst = (
            Path.home()
            / "AppData"
            / "Local"
            / "Microsoft"
            / "Office"
            / "Excel.officeUI"
        )
        office_ui_dst.parent.mkdir(parents=True, exist_ok=True)
        content = office_ui_src.read_text(encoding="utf-8")
        user_xlstart = str(
            Path.home() / "AppData" / "Roaming" / "Microsoft" / "Excel" / "XLSTART"
        )
        xlstart_personal = f'{user_xlstart}\\PERSONAL.XLSB'
        content = re.sub(
            r'onAction="[^"]*\\PERSONAL\.XLSB!',
            lambda _: f'onAction="{xlstart_personal}!',
            content,
        )
        office_ui_dst.write_text(content, encoding="utf-8")
        click.echo(f"Copied Excel.officeUI to {office_ui_dst.parent}")

        # Git Bash setup — write .bashrc alias and .minttyrc font config
        with open(Path.home() / ".bashrc", "w") as f:
            f.write(f"{BID_ALIAS} \n")
        with open(Path.home() / ".minttyrc", "w") as f:
            f.write("FontFamily=Victor Mono\nFontSize=15\n")
        click.echo("Git Bash .bashrc and .minttyrc configured.")
    elif platform.system() == "Darwin":
        # click.echo(RFQ)
        # click.echo(BID_ALIAS)
        # bid_alias = f"alias bid=\"uv run --quiet {Path(r'~/Jason Electronics Pte Ltd/Bid Proposal - Documents/@tools/bid.py').expanduser().resolve()}\""
        # click.echo(bid_alias)
        # click.echo(Path.home() / ".bash_profile")
        # click.echo(Path.home() / ".bashrc")
        pass
    else:
        pass


@click.command()
def test():
    test = Path("~/Downloads")
    open_with_default_app(test)


@click.command()
@click.argument("xl_file", default="")
@click.option(
    "-f",
    "--font",
    is_flag=True,
    help="Apply font formatting only (shows template menu, default: Arial Size 12)",
)
def beautify(xl_file: str, font: bool) -> None:
    """
    Beautify excel file.

    By default, applies smart width (based on content, max 80 chars, word wrap).
    Use -f to apply font formatting only (no smart width).
    """
    from util.beautify import beautify as _beautify

    ctx = click.Context(_beautify)
    ctx.invoke(_beautify, xl_file=xl_file, font=font)


@click.command()
@click.option("-o", "--outline", is_flag=True, help="Add outline to file from filename")
@click.option("-y", "--yes", is_flag=True, help="Answer yes to the current directory")
@click.option(
    "-t", "--toc", is_flag=True, help="Add table of contends in separate page"
)
@click.option(
    "-m",
    "--manifest",
    "use_manifest",
    is_flag=True,
    help="Use manifest file (md/txt) specifying output name and files to combine",
)
@click.option(
    "-c",
    "--create-manifest",
    "create_manifest",
    is_flag=True,
    help="Create manifest.md listing all PDFs recursively (for use with --manifest)",
)
def combine_pdf(outline: bool, toc: bool, yes: bool, use_manifest: bool, create_manifest: bool):
    """
    Combine PDFs in current folder (Alias: cpdf).

    If the combined file already exists, it will remove it first and re-combine.

    With --manifest flag, also searches for PDFs in the @docs SharePoint folder.
    """
    from util.pdfx import combine_pdf as _combine_pdf, create_manifest_file

    if create_manifest:
        create_manifest_file(Path.cwd())
        return

    # Use DOCS as fallback directory when in manifest mode
    docs_path = DOCS if use_manifest else None

    ctx = click.Context(_combine_pdf)
    ctx.invoke(
        _combine_pdf,
        outline=outline,
        toc=toc,
        yes=yes,
        use_manifest=use_manifest,
        docs_path=docs_path,
    )


@click.command()
@click.argument("directory", default="", type=click.Path())
@click.option(
    "-i",
    "--import-file",
    "import_file",
    type=click.Path(exists=True),
    help="Import from SharePoint exported CSV/Excel file",
)
@click.option(
    "-o",
    "--output",
    "output_file",
    type=click.Path(),
    help="Export report to CSV file",
)
@click.option(
    "--api",
    is_flag=True,
    help="Use m365 CLI API (requires setup and admin consent)",
)
@click.option(
    "-f",
    "--fetch",
    "fetch_url",
    type=str,
    help="Fetch from SharePoint URL using browser automation",
)
@click.option(
    "-l",
    "--library",
    default="Documents",
    help="SharePoint library title for --fetch (default: Documents)",
)
@click.option(
    "--folder",
    default="",
    help="Folder path within library for --fetch (e.g., '@docs')",
)
@click.option(
    "--debug",
    is_flag=True,
    help="Debug mode: show browser window for --fetch",
)
@click.option(
    "--month",
    "filter_month",
    type=str,
    default=None,
    help="Filter by month: YYYY-MM or range YYYY-MM:YYYY-MM. Default: current month",
)
@click.option(
    "--all-time",
    "all_time",
    is_flag=True,
    help="Show all data without time filtering",
)
@click.option(
    "--person",
    "filter_person",
    type=str,
    default=None,
    help="Filter to specific person. Use 'all' for everyone",
)
def audit(
    directory: str,
    import_file: str | None,
    output_file: str | None,
    api: bool,
    fetch_url: str | None,
    library: str,
    folder: str,
    debug: bool,
    filter_month: str | None,
    all_time: bool,
    filter_person: str | None,
):
    """
    Audit folder to track file contributions.

    \b
    Default audits @docs folder. For full author info, use --fetch or --import-file.

    \b
    Examples:
      bid audit -f URL --folder "@docs"                    # Current month, interactive person
      bid audit -f URL --folder "@docs" --person all       # Current month, all people
      bid audit -f URL --folder "@docs" --all-time         # All time, interactive person
      bid audit -f URL --folder "@docs" --month 2025-12    # Specific month
    """
    from util.audit import audit as _audit

    ctx = click.Context(_audit)
    ctx.invoke(
        _audit,
        directory=directory,
        import_file=import_file,
        output_file=output_file,
        api=api,
        fetch_url=fetch_url,
        library=library,
        folder=folder,
        debug=debug,
        filter_month=filter_month,
        all_time=all_time,
        filter_person=filter_person,
    )


@click.command()
@click.argument("folder_name", default="")
def vo(folder_name: str) -> None:
    """
    Create VO (Variation Order) folder structure under an existing project.

    Searches for the project folder in @rfqs and creates a new VO subfolder
    with the standard folder structure.
    """
    # Handle case where @rfqs does not exist
    if not Path(RFQ).expanduser().exists():
        click.echo("The folder @rfqs does not exist. Check if you have access.")
        return

    # Get project folder name/job code
    if folder_name == "":
        folder_name = click.prompt("Please enter project folder name or job code")

    # Search for matching folders
    matches = find_project_folder(folder_name)

    if not matches:
        click.echo(f"No project folder found matching '{folder_name}'")
        return

    # Handle multiple matches
    if len(matches) == 1:
        project_path, year = matches[0]
    else:
        click.echo("Multiple folders found:")
        for i, (path, year) in enumerate(matches, 1):
            click.echo(f"  {i}. {path.name} ({year})")

        choice = click.prompt(
            "Select folder number",
            type=click.IntRange(1, len(matches)),
        )
        project_path, year = matches[choice - 1]

    # Confirm with user
    click.echo(f"Found: {project_path.name} ({year})")
    if not click.confirm("Use this folder?", default=True):
        click.echo("Aborted.")
        return

    # Find or create 07-VO folder
    vo_parent = project_path / "07-VO"
    vo_parent.mkdir(exist_ok=True)

    # Clean up legacy empty 00-Arc folder if exists
    legacy_arc = vo_parent / "00-Arc"
    if legacy_arc.exists() and legacy_arc.is_dir() and not any(legacy_arc.iterdir()):
        try:
            legacy_arc.rmdir()
            click.echo("Removed empty legacy folder: 07-VO/00-Arc")
        except OSError as e:
            click.echo(f"Warning: could not remove legacy folder 07-VO/00-Arc: {e}")

    # List existing VO folders
    existing_vos = sorted(
        [
            f.name
            for f in vo_parent.iterdir()
            if f.is_dir() and re.match(r"^\d{2}-VO\b", f.name)
        ]
    )
    if existing_vos:
        click.echo("Existing VO folders:")
        for vo_name_existing in existing_vos:
            click.echo(f"  {vo_name_existing}")
    else:
        click.echo("No existing VO folders.")

    # Get next VO number
    vo_number = get_next_vo_number(vo_parent)

    # Get VO name with explanation
    click.echo(f"New folder will be named: {vo_number:02d}-VO <your input>")
    vo_name = click.prompt("Please enter VO name")
    vo_name = vo_name.strip().upper()

    vo_folder_name = f"{vo_number:02d}-VO {vo_name}"
    vo_path = vo_parent / vo_folder_name

    if vo_path.exists():
        click.echo(f"Folder '{vo_folder_name}' already exists.")
        return

    # Create the VO folder and structure
    click.echo(f"Creating: 07-VO/{vo_folder_name}")
    vo_path.mkdir()
    create_vo_structure(vo_path)
    click.echo("Created folder structure successfully.")

    # Create new commercial Proposal
    if click.confirm("Do you want to create a new Commercial Proposal?"):
        from util import excelx

        version = click.prompt("Version No, default:", default="R0", show_default=True)
        version = str(version).upper()

        # VO number prefix (e.g., "V1" for 01-VO)
        vo_prefix = f"V{vo_number}"

        template_folder = (
            Path(RFQ).expanduser().parent.absolute().resolve() / "@tools/resources"
        )
        template = "Template.xlsx"
        commercial_folder = vo_path / "01-Commercial"

        # Filename: "<project_name> <vo_name> V1-R0.xlsx"
        commercial_file = f"{project_path.name} {vo_name} {vo_prefix}-{version}.xlsx"

        # Jobcode: "J12789 V1"
        jobcode = f"{project_path.name.split()[0]} {vo_prefix}"

        excelx.create_excel_from_template(
            template_folder,
            template,
            commercial_folder,
            commercial_file,
            jobcode,
            version,
        )

    # Ask to open folder
    if click.confirm("Do you want to open the VO folder?", default=True):
        open_with_default_app(vo_path)


@click.command()
@click.argument("folder_name", default="")
def ho(folder_name: str) -> None:
    """
    Handover project folders from @rfqs to @handover.

    Syncs project or VO folders for handover after successful bidding.
    """
    # Handle case where @rfqs or @handover does not exist
    if not Path(RFQ).expanduser().exists():
        click.echo("The folder @rfqs does not exist. Check if you have access.")
        return

    if not Path(HO).expanduser().exists():
        click.echo("The folder @handover does not exist. Check if you have access.")
        return

    # Get project folder name/job code
    if folder_name == "":
        folder_name = click.prompt("Please enter project folder name or job code")

    # Search for matching folders
    matches = find_project_folder(folder_name)

    if not matches:
        click.echo(f"No project folder found matching '{folder_name}'")
        return

    # Handle multiple matches
    if len(matches) == 1:
        project_path, year = matches[0]
    else:
        click.echo("Multiple folders found:")
        for i, (path, year) in enumerate(matches, 1):
            click.echo(f"  {i}. {path.name} ({year})")

        choice = click.prompt(
            "Select folder number",
            type=click.IntRange(1, len(matches)),
        )
        project_path, year = matches[choice - 1]

    # Confirm with user
    click.echo(f"Found: {project_path.name} ({year})")
    if not click.confirm("Use this folder?", default=True):
        click.echo("Aborted.")
        return

    # Get handover candidates
    candidates = get_handover_candidates(project_path)

    if len(candidates) == 1:
        # Only main folder available
        selected_name, source_path = candidates[0]
        click.echo("Handover candidate: 00-MAIN (Main Project)")
        if not click.confirm("Proceed with this folder?", default=True):
            click.echo("Aborted.")
            return
    else:
        click.echo("Handover candidates:")
        for i, (name, _) in enumerate(candidates, 1):
            label = "(Main Project)" if name == "00-MAIN" else ""
            click.echo(f"  {i}. {name} {label}")

        choice = click.prompt(
            "Select folder to handover",
            type=click.IntRange(1, len(candidates)),
        )
        selected_name, source_path = candidates[choice - 1]

    # Determine if this is a VO handover
    is_vo = selected_name != "00-MAIN"

    # Prepare destination in @handover
    handover_root = Path(HO).expanduser()
    project_dest = handover_root / project_path.name
    dest_path = project_dest / selected_name

    # Check for legacy folder structure
    if dest_path.exists() and not is_conforming_handover_folder(dest_path):
        click.echo("\nWarning: Destination folder exists with legacy structure:")
        click.echo(f"  {dest_path}")
        click.echo("Syncing may override existing content.")
        if not click.confirm("Do you want to continue?", default=False):
            click.echo("Aborted.")
            return

    click.echo(f"\nSyncing to: {dest_path}")

    # Perform the sync
    perform_handover_sync(source_path, dest_path, is_vo)

    click.echo("\nHandover sync complete.")
    click.echo("NOTE: Please update 05-Cost folder manually.")

    # Ask to open folder
    if click.confirm("Do you want to open the handover folder?", default=True):
        open_with_default_app(dest_path)


@click.command()
@click.argument("folder_name", default="")
def co(folder_name: str) -> None:
    """
    Create costing folder structure in @costing for a project.

    Creates folder for manual placement of costing files after successful bidding.
    """
    # Handle case where @rfqs or @costing does not exist
    if not Path(RFQ).expanduser().exists():
        click.echo("The folder @rfqs does not exist. Check if you have access.")
        return

    if not Path(CO).expanduser().exists():
        click.echo("The folder @costing does not exist. Check if you have access.")
        return

    # Get project folder name/job code
    if folder_name == "":
        folder_name = click.prompt("Please enter project folder name or job code")

    # Search for matching folders
    matches = find_project_folder(folder_name)

    if not matches:
        click.echo(f"No project folder found matching '{folder_name}'")
        return

    # Handle multiple matches
    if len(matches) == 1:
        project_path, year = matches[0]
    else:
        click.echo("Multiple folders found:")
        for i, (path, year) in enumerate(matches, 1):
            click.echo(f"  {i}. {path.name} ({year})")

        choice = click.prompt(
            "Select folder number",
            type=click.IntRange(1, len(matches)),
        )
        project_path, year = matches[choice - 1]

    # Confirm with user
    click.echo(f"Found: {project_path.name} ({year})")
    if not click.confirm("Use this folder?", default=True):
        click.echo("Aborted.")
        return

    # Get costing candidates (same as handover candidates)
    candidates = get_handover_candidates(project_path)

    if len(candidates) == 1:
        # Only main folder available
        selected_name, _ = candidates[0]
        click.echo("Costing candidate: 00-MAIN (Main Project)")
        if not click.confirm("Proceed with this folder?", default=True):
            click.echo("Aborted.")
            return
    else:
        click.echo("Costing candidates:")
        for i, (name, _) in enumerate(candidates, 1):
            label = "(Main Project)" if name == "00-MAIN" else ""
            click.echo(f"  {i}. {name} {label}")

        choice = click.prompt(
            "Select folder for costing",
            type=click.IntRange(1, len(candidates)),
        )
        selected_name, _ = candidates[choice - 1]

    # Prepare destination in @costing
    costing_root = Path(CO).expanduser()
    project_dest = costing_root / project_path.name
    dest_path = project_dest / selected_name

    # Create the folder
    if dest_path.exists():
        click.echo(f"Folder already exists: {dest_path}")
    else:
        dest_path.mkdir(parents=True, exist_ok=True)
        click.echo(f"\nCreated: {dest_path}")

    click.echo("Please copy costing files manually to this folder.")

    # Ask to open folder
    if click.confirm("Do you want to open the costing folder?", default=True):
        open_with_default_app(dest_path)


# ── Rate calculation helpers ──────────────────────────────────────────────────

def _ceil_to(value: float, n: int) -> int:
    return math.ceil(value / n) * n


def _round_to(value: float, n: int) -> int:
    return round(value / n) * n


def _fmt_rate(value: object) -> str:
    if isinstance(value, str):
        return value
    return f"{int(value):,}"


def _calc_standard(day: float, mode: str) -> dict:
    # Formulas verified against 6 data points (MODEC 2025, MODEC 2026, HOS 2026).
    # OT = ceil_to_5(day / 7.5)  ← 4/3 × of 10-hr hourly rate
    # Standby = round_to_25(day × 0.73)
    ot = _ceil_to(day / 7.5, 5)
    standby = _round_to(day * 0.73, 25)
    sun_ph: int | str = ot if mode == "onshore" else "N/A"
    return {"day": int(day), "ot": ot, "sun_ph": sun_ph, "standby": standby}


def _calc_seatrium(day: float, mode: str) -> dict:
    # Calibrated from Seatrium 2026 rates.
    # OT = ceil_to_5(day × 0.15)  ← 1.5 × of 10-hr hourly rate
    # Onshore standby: 80% (SGD basis)
    ot = _ceil_to(day * 0.15, 5)
    if mode == "onshore":
        sun_ph_day = _ceil_to(day * 1.5, 50)
        sun_ph_ot = _ceil_to(sun_ph_day * 0.15, 5)
        standby = _round_to(day * 0.80, 5)
        return {
            "day": int(day), "ot": ot,
            "sun_ph_day": sun_ph_day, "sun_ph_ot": sun_ph_ot,
            "standby": standby,
        }
    else:
        standby = _round_to(day * 0.82, 50)
        return {"day": int(day), "ot": ot, "standby": standby}


def _print_rate_rows(rows: list[tuple[str, object, str | None]], indent: int = 4) -> None:
    pad = " " * indent
    w = max(len(k) for k, *_ in rows)
    for key, val, note in rows:
        note_str = f"  ({note})" if note else ""
        click.echo(f"{pad}{key:<{w}}  {_fmt_rate(val):>8}{note_str}")


def _print_std_section(mode: str, rates: dict, designation: str) -> None:
    label = "ONSHORE" if mode == "onshore" else "OFFSHORE"
    hours = 10 if mode == "onshore" else 12
    desc = f"Mon-Sat, {hours} hrs/day" if mode == "onshore" else f"Mon-Sun, {hours} hrs/day"
    hourly = f"{round(rates['day'] / hours, 1):g}/hr"
    click.echo(f"{label}  ·  {designation}  ({desc})")
    click.echo()
    _print_rate_rows([
        ("Day Rate",  rates["day"],     hourly),
        ("OT/hr",     rates["ot"],      "×4/3"),
        ("Sun/PH/hr", rates["sun_ph"],  None),
        ("Standby",   rates["standby"], None),
    ])
    click.echo()


def _print_seatrium_section(mode: str, rates: dict, designation: str, usd_rate: float | None = None) -> None:
    label = "ONSHORE" if mode == "onshore" else "OFFSHORE"
    hours = 10 if mode == "onshore" else 12
    desc = f"Mon-Sat, {hours} hrs/day" if mode == "onshore" else f"Mon-Sun, {hours} hrs/day"
    hourly = f"{round(rates['day'] / hours, 1):g}/hr"
    click.echo(f"{label}  ·  {designation}  ({desc})")
    click.echo()
    if mode == "onshore":
        ph_hourly = f"{_ceil_to(rates['sun_ph_day'] / hours, 5)}/hr"
        rows: list[tuple[str, object, str | None]] = [
            ("Day Rate",          rates["day"],         hourly),
            ("OT/hr",             rates["ot"],          "×3/2"),
            ("Sun/PH Day Rate",   rates["sun_ph_day"],  ph_hourly),
            ("Sun/PH OT/hr",      rates["sun_ph_ot"],   "×3/2"),
            ("Standby",           rates["standby"],     None),
        ]
    else:
        rows = [
            ("Day Rate", rates["day"],     hourly),
            ("OT/hr",    rates["ot"],      "×3/2"),
            ("Standby",  rates["standby"], None),
        ]
    if usd_rate:
        w = max(len(r[0]) for r in rows)
        click.echo(f"    {'':>{w}}  {'SGD':>8}   {'USD':>8}")
        for lbl, sgd, note in rows:
            usd = _ceil_to(sgd / usd_rate, 100)
            note_str = f"  ({note})" if note else ""
            click.echo(f"    {lbl:<{w}}  {_fmt_rate(sgd):>8}   {_fmt_rate(usd):>8}{note_str}")
    else:
        _print_rate_rows(rows)
    click.echo()



@click.command("rate")
@click.option("--onshore", "onshore_rate", type=float, default=None, metavar="RATE",
              help="Onshore day rate (Mon-Sat, 10 hrs/day)")
@click.option("--offshore", "offshore_rate", type=float, default=None, metavar="RATE",
              help="Offshore day rate (Mon-Sun, 12 hrs/day)")
@click.option("--specialist", "specialist_tier",
              type=click.Choice(["standard", "premium", "super"], case_sensitive=False),
              default=None,
              help="Specialist tier rate card (standard|premium|super)")
@click.option("--tiers", "show_tiers", is_flag=True, default=False,
              help="Show specialist tier definitions in SGD/USD/EUR/GBP")
@click.option("--special", is_flag=True,
              help="MODEC/HOS rate structure, ×4/3 OT (default: ×3/2 standard)")
def rate_cmd(
    onshore_rate: float | None,
    offshore_rate: float | None,
    specialist_tier: str | None,
    show_tiers: bool,
    special: bool,
) -> None:
    """Calculate OT, Sun/PH, and Standby from a man-day rate.

    \b
    Standard (×3/2 OT):
      bid rate --onshore 1100
      bid rate --offshore 1650
      bid rate --onshore 1100 --offshore 1650

    \b
    MODEC/HOS special rates (×4/3 OT):
      bid rate --onshore 875 --special
      bid rate --offshore 1700 --special

    \b
    Specialist tier overview (thresholds in SGD/USD/EUR/GBP):
      bid rate --tiers

    \b
    Specialist tier rate card (specific tier only):
      bid rate --specialist standard
      bid rate --specialist premium
      bid rate --specialist super

    \b
    Specialist with custom selling rate:
      bid rate --onshore 3500 --specialist super
    """
    if show_tiers:
        cfg = _load_mob_config()
        spec = cfg.get("specialist", {})
        tiers = spec["tiers"]
        fx = spec.get("fx", {})
        currencies = ["SGD", "USD", "EUR", "GBP"]
        tier_names = ("standard", "premium", "super")
        w_tier = 10
        w_col = 10
        click.echo("SPECIALIST TIERS  ·  Supplier Onshore Day Rate Threshold  (cost figures)")
        click.echo()
        header = f"  {'Tier':<{w_tier}}" + "".join(f"  {c:>{w_col}}" for c in currencies)
        click.echo(header)
        click.echo("  " + "─" * (w_tier + (w_col + 2) * len(currencies)))
        def _thresh(usd_val: float, currency: str) -> float:
            sgd_val = usd_val / fx.get("USD", 1)
            raw = sgd_val * fx.get(currency, 1) if currency != "SGD" else sgd_val
            return math.ceil(raw / 100) * 100

        for t_name in tier_names:
            tier = tiers[t_name]
            usd_max = tier["day_rate_usd_max"]
            if usd_max >= 99999:
                ref = tiers["premium"]["day_rate_usd_max"]
                vals = {c: f"> {_thresh(ref, c):,.0f}" for c in currencies}
            else:
                vals = {c: f"≤ {_thresh(usd_max, c):,.0f}" for c in currencies}
            row = f"  {t_name.title():<{w_tier}}" + "".join(f"  {vals[c]:>{w_col}}" for c in currencies)
            click.echo(row)
        click.echo()
        return

    if specialist_tier:
        cfg = _load_mob_config()
        spec = cfg.get("specialist", {})
        tiers = spec["tiers"]
        short_haul = spec.get("short_haul_countries", [])
        has_custom_rate = onshore_rate is not None or offshore_rate is not None

        usd_rate = cfg["defaults"]["usd_exchange_rate"]
        usd_round = cfg["defaults"]["usd_round"]

        def _print_spec_mob(t_name: str, tier: dict) -> None:
            click.echo(f"SELLING PRICE  ·  MOB/DEMOB  ·  Specialist  ·  {t_name.title()}")
            click.echo()
            w = len("Mob/Demob")
            for haul, mob_sgd, note in [
                ("Short-haul", tier["mob_short"], ", ".join(short_haul)),
                ("Long-haul",  tier["mob_long"],  "all other destinations"),
            ]:
                mob_usd = _ceil_to(mob_sgd / usd_rate, usd_round)
                click.echo(f"    {haul}  ({note})")
                click.echo(f"    {'':>{w}}  {'SGD':>8}   {'USD':>8}")
                for lbl, sgd, usd in [("Mob", mob_sgd, mob_usd), ("Demob", mob_sgd, mob_usd), ("Mob/Demob", mob_sgd * 2, mob_usd * 2)]:
                    click.echo(f"    {lbl:<{w}}  {_fmt_rate(sgd):>8}   {_fmt_rate(usd):>8}")
                click.echo()

        tier = tiers[specialist_tier]
        label = f"Specialist  ·  {specialist_tier.title()}"
        usd_rate = cfg["defaults"]["usd_exchange_rate"]
        onshore_sgd = onshore_rate if onshore_rate is not None else _ceil_to(tier["onshore_sell"] * usd_rate, 100)
        offshore_sgd = offshore_rate if offshore_rate is not None else _ceil_to(tier["offshore_sell"] * usd_rate, 100)
        click.echo("SELLING PRICE  (SGD / USD)")
        click.echo()
        for mode, sell in [("onshore", onshore_sgd), ("offshore", offshore_sgd)]:
            _print_seatrium_section(mode, _calc_seatrium(sell, mode), label, usd_rate=usd_rate)
        _print_spec_mob(specialist_tier, tier)
        return

    has_rates = onshore_rate is not None or offshore_rate is not None
    if not has_rates:
        raise click.UsageError("Provide --onshore RATE and/or --offshore RATE, or --specialist TIER.")

    for mode, day in [("onshore", onshore_rate), ("offshore", offshore_rate)]:
        if day is None:
            continue
        if special:
            _print_std_section(mode, _calc_standard(day, mode), "JEN Engineer")
        else:
            _print_seatrium_section(mode, _calc_seatrium(day, mode), "JEN Engineer")


# ── End rate helpers ──────────────────────────────────────────────────────────


# ── Mob/demob helpers ─────────────────────────────────────────────────────────

def _load_mob_config() -> dict:
    config_path = Path(__file__).parent / "util" / "mob_config.toml"
    if not config_path.exists():
        raise click.ClickException(f"Mob config not found: {config_path}")
    with open(config_path, "rb") as f:
        return tomllib.load(f)


@click.command("mob")
@click.argument("country", required=False, metavar="COUNTRY")
@click.option("--batam", is_flag=True, help="Batam deployment (ferry, no visa)")
@click.option("--offshore", is_flag=True, help="Offshore deployment (2 hotel nights only; default: onshore)")
@click.option("--days", "days_override", default=None, type=int, metavar="DAYS",
              help="Override mob period in working days")
@click.option("--buffer", "buffers", nargs=2, multiple=True, metavar="AMOUNT LABEL",
              help="Additional cost item, repeatable (e.g. --buffer 1000 'agent fee')")
@click.option("--specialist", "specialist_tier",
              type=click.Choice(["standard", "premium", "super"], case_sensitive=False),
              default=None,
              help="Specialist mob by tier")
@click.option("--day-rate", "spec_day_rate", type=float, default=None, metavar="RATE",
              help="Supplier onshore day rate to auto-classify tier; offshore rates are not used for classification (use with --currency)")
@click.option("--currency", "spec_currency",
              type=click.Choice(["USD", "SGD", "EUR", "GBP", "NOK", "DKK"], case_sensitive=False),
              default="USD", show_default=True,
              help="Currency of --day-rate")
def mob_cmd(
    country: str | None,
    batam: bool,
    offshore: bool,
    days_override: int | None,
    buffers: tuple,
    specialist_tier: str | None,
    spec_day_rate: float | None,
    spec_currency: str,
) -> None:
    """Estimate mob/demob lumpsum for engineer or specialist deployment.

    \b
    JEN engineer (SGD, from Singapore):
      bid mob BR
      bid mob BR --offshore
      bid mob --batam
      bid mob US --buffer 1000 "agent fee" --buffer 500 "medical"

    \b
    Specialist (USD, tier-based selling price):
      bid mob --specialist standard
      bid mob --specialist super NA
      bid mob --day-rate 5101 --currency SGD NA
    """
    # ── Specialist path ───────────────────────────────────────────────────────
    if specialist_tier or spec_day_rate is not None:
        cfg = _load_mob_config()
        spec = cfg.get("specialist", {})
        tiers = spec.get("tiers", {})

        if spec_day_rate is not None:
            fx = spec.get("fx", {})
            usd_rate = cfg["defaults"]["usd_exchange_rate"]
            rate_sgd = spec_day_rate / fx.get(spec_currency.upper(), 1.0)
            rate_usd = rate_sgd / usd_rate
            auto_tier = next(
                n for n in ("standard", "premium", "super")
                if rate_usd <= tiers[n]["day_rate_usd_max"]
            )
            if specialist_tier and specialist_tier != auto_tier:
                click.echo(
                    f"  Note: {spec_currency.upper()} {_fmt_rate(spec_day_rate)}/day"
                    f" → SGD {round(rate_sgd):,}/day"
                    f" → USD {round(rate_usd):,}/day classifies as {auto_tier};"
                    f" using specified {specialist_tier} instead.", err=True
                )
            else:
                specialist_tier = auto_tier
                click.echo(
                    f"  {spec_currency.upper()} {_fmt_rate(spec_day_rate)}/day"
                    f"  →  SGD {round(rate_sgd):,}/day"
                    f"  →  USD {round(rate_usd):,}/day"
                    f"  →  {specialist_tier.title()}"
                )
                click.echo()

        tier = tiers[specialist_tier]
        short_haul = spec.get("short_haul_countries", [])

        if country:
            code = country.upper()
            is_long = code not in short_haul
            haul = "Long-haul" if is_long else "Short-haul"
            mob = tier["mob_long"] if is_long else tier["mob_short"]
            dest = cfg.get("countries", {}).get(code, {}).get("name", code)
            usd_rate = cfg["defaults"]["usd_exchange_rate"]
            usd_round = cfg["defaults"]["usd_round"]
            mob_usd = _ceil_to(mob / usd_rate, usd_round)
            w = len("Mob/Demob")
            click.echo(f"SELLING PRICE  ·  SPECIALIST  ·  {specialist_tier.title()}  ·  {dest}  ({haul})")
            click.echo()
            click.echo(f"    {'':>{w}}  {'SGD':>8}   {'USD':>8}")
            for lbl, sgd, usd in [("Mob", mob, mob_usd), ("Demob", mob, mob_usd), ("Mob/Demob", mob * 2, mob_usd * 2)]:
                click.echo(f"    {lbl:<{w}}  {_fmt_rate(sgd):>8}   {_fmt_rate(usd):>8}")
        else:
            usd_rate = cfg["defaults"]["usd_exchange_rate"]
            usd_round = cfg["defaults"]["usd_round"]
            w = len("Mob/Demob")
            click.echo(f"SELLING PRICE  ·  SPECIALIST  ·  {specialist_tier.title()}")
            click.echo()
            for haul, mob_sgd, note in [
                ("Short-haul", tier["mob_short"], ", ".join(short_haul)),
                ("Long-haul",  tier["mob_long"],  "all other destinations"),
            ]:
                mob_usd = _ceil_to(mob_sgd / usd_rate, usd_round)
                click.echo(f"    {haul}  ({note})")
                click.echo(f"    {'':>{w}}  {'SGD':>8}   {'USD':>8}")
                for lbl, sgd, usd in [("Mob", mob_sgd, mob_usd), ("Demob", mob_sgd, mob_usd), ("Mob/Demob", mob_sgd * 2, mob_usd * 2)]:
                    click.echo(f"    {lbl:<{w}}  {_fmt_rate(sgd):>8}   {_fmt_rate(usd):>8}")
                click.echo()

        click.echo()
        click.echo(f"  → Rate card: bid rate --specialist {specialist_tier}")
        click.echo()
        return
    # ── End specialist path ───────────────────────────────────────────────────

    if not batam and not country:
        raise click.UsageError("Provide COUNTRY or --batam.")
    if batam and country:
        raise click.UsageError("Cannot use both COUNTRY and --batam.")

    cfg = _load_mob_config()
    defaults = cfg["defaults"]
    buf_pct = defaults["airfare_buffer_pct"] / 100
    change_fee = defaults["airfare_change_fee"]
    travel_day_rate = defaults["engineer_travel_day_rate"]
    sg_transport_ow = defaults["sg_transport_one_way"]
    allowance_per_day = defaults["allowance_per_day"]
    buc_gm = defaults["buc_gm_pct"] / 100
    selling_gm = defaults["selling_gm_pct"] / 100
    usd_rate = defaults["usd_exchange_rate"]
    usd_round = defaults["usd_round"]
    lumpsum_round = defaults["lumpsum_round"]
    std_days = defaults["mob_days"]

    if offshore and days_override is not None:
        click.echo("Note: --days is ignored for offshore; standard rate applies.", err=True)

    if batam:
        data = cfg["batam"]
        label = data["name"]
        routing = False
        country_extra = 0
        working_days = std_days if offshore else (days_override if days_override is not None else std_days)
        days = working_days
        fare_raw = data["ferry_one_way"]
        total_fare = _ceil_to(fare_raw * 2 * (1 + buf_pct), 5)
        fare_key = "Ferry (2× one-way)"
        visa = data.get("visa", 0)
        visa_note: str | None = None
    else:
        code = country.upper()  # type: ignore[union-attr]
        countries = cfg.get("countries", {})
        if code not in countries:
            available = ", ".join(sorted(countries.keys()))
            raise click.ClickException(
                f"Country '{code}' not in config. Available: {available}"
            )
        data = countries[code]
        label = data["name"]
        routing = data.get("routing", False)
        if offshore:
            working_days = std_days
            routing_extra = 0
            country_extra = 0
        else:
            working_days = days_override if days_override is not None else std_days
            routing_extra = defaults["routing_extra_days"] if routing and days_override is None else 0
            country_extra = data.get("extra_days", 0) if days_override is None else 0
        days = working_days + routing_extra + country_extra
        fare_raw = data["airfare_roundtrip"]
        total_fare = _ceil_to(fare_raw * (1 + buf_pct), 50)
        fare_key = "Airfare (RT)"
        visa = data.get("visa", 0)
        visa_note = data.get("visa_note")

    transport = (data.get("transport_one_way", 0) + sg_transport_ow) * 2
    allowance = allowance_per_day * days
    travel_days = data.get("travel_days_one_way", 1) * 2
    travel_cost = travel_days * travel_day_rate
    extra_total = sum(int(float(amt)) for amt, _ in buffers)
    applied_change_fee = 0 if (batam or fare_raw == 0) else change_fee
    hotel_rate = data.get("hotel_per_night", 0)
    hotel = hotel_rate * working_days
    subtotal = total_fare + applied_change_fee + visa + transport + travel_cost + allowance + hotel + extra_total
    lumpsum = _ceil_to(subtotal, lumpsum_round)

    tags = []
    if not offshore and not batam and days_override is None:
        if routing:
            tags.append(f"routing +{defaults['routing_extra_days']}d")
        if country_extra:
            tags.append(f"transit +{country_extra}d")
    tag_str = f"  ({', '.join(tags)})" if tags else ""

    if not offshore and not batam and data.get("onshore_mob_absorbed"):
        click.echo(f"{label.upper()}  ·  Onshore  ·  SGD")
        click.echo()
        click.echo("  Mob/demob not applicable — cost absorbed in engineer day rates.")
        return

    if offshore:
        mode_label = "Offshore (anchorage / sea trial)"
    elif days_override is not None:
        mode_label = "Onshore"
    else:
        mode_label = "Onshore / Offshore"
    header_parts = [label.upper()]
    if mode_label:
        header_parts.append(mode_label)
    header_parts += [f"{days} days{tag_str}", "SGD"]
    click.echo("  ·  ".join(header_parts))

    fare_note = (f"mob + demob tickets, +{defaults['airfare_buffer_pct']}% buffer" if batam
                 else f"flexible return, +{defaults['airfare_buffer_pct']}% buffer")
    pad = "  "

    rows: list[tuple[str, object, str | None]] = [
        (fare_key,          total_fare,          fare_note),
        ("Date Change Fee", applied_change_fee,  "return date change" if applied_change_fee else None),
        ("Work Visa",       visa,                f"⚠  {visa_note}" if visa_note else None),
        ("Transport",       transport,           "SG + destination, arrival + departure"),
        ("Travel Time",     travel_cost,         f"500/day × {travel_days} days" if travel_cost else None),
        ("Allowance",       allowance,           f"100/day × {days} days"),
        ("Hotel",           hotel,               f"{hotel_rate}/night × {working_days} nights"),
    ]
    for amt, lbl in buffers:
        rows.append((lbl, int(float(amt)), "manual"))

    display_rows = [(k, v, n) for k, v, n in rows if v]
    w = max((len(k) for k, *_ in display_rows), default=14)
    w = max(w, len("Lumpsum (SGD)"), len("Base Unit Cost"))

    click.echo()
    _print_rate_rows(display_rows, indent=4)
    click.echo()

    sep = "─" * (w + 12)
    click.echo(f"    {sep}")
    click.echo(f"    {'Subtotal':<{w}}  {_fmt_rate(subtotal):>8}")
    click.echo(f"    {'Lumpsum (SGD)':<{w}}  {_fmt_rate(lumpsum):>8}")

    buc     = _ceil_to(lumpsum / (1 - buc_gm), lumpsum_round)
    sp_sgd  = _ceil_to(buc / (1 - selling_gm), lumpsum_round)
    mob_sgd = _round_to(sp_sgd / 2, lumpsum_round)
    mob_usd = _ceil_to(mob_sgd / usd_rate, usd_round)

    click.echo()
    click.echo(f"    COST")
    click.echo(f"    {'Base Unit Cost':<{w}}  {_fmt_rate(buc):>8}  SGD  (+{int(buc_gm*100)}% GM)")
    click.echo()
    click.echo(f"    SELLING PRICE  (+{int(selling_gm*100)}% GM on BUC)")
    click.echo(f"    {'':>{w}}  {'SGD':>8}   {'USD':>8}")
    click.echo(f"    {'Mob':<{w}}  {_fmt_rate(mob_sgd):>8}   {_fmt_rate(mob_usd):>8}")
    click.echo(f"    {'Demob':<{w}}  {_fmt_rate(mob_sgd):>8}   {_fmt_rate(mob_usd):>8}")
    click.echo(f"    {'Mob/Demob':<{w}}  {_fmt_rate(mob_sgd * 2):>8}   {_fmt_rate(mob_usd * 2):>8}")


@click.command("mob-config")
@click.argument("output", default="mob_config.xlsx", required=False, metavar="OUTPUT")
def mob_config_cmd(output: str) -> None:
    """Export mob/demob config to Excel for review and sharing.

    \b
    Examples:
      bid mob-config
      bid mob-config ~/Desktop/mob_config.xlsx
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    cfg = _load_mob_config()
    defaults = cfg["defaults"]
    wb = openpyxl.Workbook()

    HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    SUBHDR_FILL = PatternFill("solid", fgColor="BDD7EE")
    SUBHDR_FONT = Font(bold=True, size=11)
    CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    WRAP      = Alignment(wrap_text=True, vertical="top")

    def _style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = CENTER

    def _autowidth(ws, extra=4, cap=45):
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + extra, cap)

    # ── Sheet 1: Countries ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Countries"
    ws.row_dimensions[1].height = 30

    headers = [
        "Code", "Name",
        "Airfare RT\n(SGD)", "Visa\n(SGD)", "Visa Note",
        "Transport\nOne-Way (SGD)", "Hotel/Night\n(SGD)", "Travel Days\n(One-Way)",
        "Routing\n(+2 days)", "Extra Days\n(transit)",
        "Comments",
    ]
    ws.append(headers)
    _style_header(ws)

    countries = cfg.get("countries", {})
    for code, data in sorted(countries.items()):
        row = [
            code,
            data.get("name", ""),
            data.get("airfare_roundtrip", 0),
            data.get("visa", 0),
            data.get("visa_note", ""),
            data.get("transport_one_way", 0),
            data.get("hotel_per_night", 0),
            data.get("travel_days_one_way", 0),
            "Yes" if data.get("routing") else "",
            data.get("extra_days", "") or "",
            "",
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = WRAP

    # Number format for SGD columns
    for col_idx in [3, 4, 6, 7]:
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = '#,##0'

    _autowidth(ws)
    ws.column_dimensions["E"].width = 35   # Visa Note
    ws.column_dimensions["K"].width = 30   # Comments
    ws.freeze_panes = "A2"

    # ── Sheet 2: Batam ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Batam")
    ws2.row_dimensions[1].height = 25
    ws2.append(["Field", "Value (SGD)", "Notes", "Comments"])
    _style_header(ws2)

    batam = cfg.get("batam", {})
    batam_notes = {
        "name":              "Display name",
        "ferry_one_way":     "One-way ferry ticket estimate (×2 for mob + demob, +20% buffer applied)",
        "visa":              "Visa cost",
        "transport_one_way": "Land transport one way (destination side)",
        "hotel_per_night":   "Hotel per night",
        "travel_days_one_way": "0 — 45-min ferry, not a full travel day",
    }
    for key, val in batam.items():
        ws2.append([key, val, batam_notes.get(key, ""), ""])
    _autowidth(ws2)
    ws2.column_dimensions["C"].width = 50
    ws2.column_dimensions["D"].width = 30

    # ── Sheet 3: Defaults ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Defaults")
    ws3.row_dimensions[1].height = 25
    ws3.append(["Setting", "Value", "Description", "Comments"])
    _style_header(ws3)

    default_desc = {
        "mob_days":               "Standard working-day mob/demob period",
        "airfare_buffer_pct":     "% buffer applied to airfare for price volatility",
        "airfare_change_fee":     "Return date change fee (SGD); assumed to occur once",
        "allowance_per_day":      "Daily engineer abroad allowance (SGD)",
        "engineer_travel_day_rate": "Engineer time cost per calendar travel day, capped at 8 hrs (SGD)",
        "sg_transport_one_way":   "Taxi to/from Changi Airport (SGD), same for all destinations",
        "lumpsum_round":          "Round final lumpsum to nearest N (SGD)",
        "routing_extra_days":     "Extra days added for routing-heavy destinations (BR, GY, NA, SN)",
        "buc_gm_pct":             "% GM applied to lumpsum to get Base Unit Cost",
        "selling_gm_pct":         "% GM applied to BUC to get selling price",
        "usd_exchange_rate":      "SGD per 1 USD exchange rate assumption",
        "usd_round":              "USD rounding granularity (finer than SGD to stay near exchange rate)",
    }
    for key, val in defaults.items():
        ws3.append([key, val, default_desc.get(key, ""), ""])
    _autowidth(ws3)
    ws3.column_dimensions["C"].width = 55
    ws3.column_dimensions["D"].width = 30

    # Shade alternate rows on all sheets for readability
    ALT_FILL = PatternFill("solid", fgColor="F2F7FC")
    for sheet in [ws, ws2, ws3]:
        for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if i % 2 == 0:
                for cell in row:
                    if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                        cell.fill = ALT_FILL

    wb.save(output)
    click.echo(f"Exported → {output}")


# ── End mob/demob helpers ─────────────────────────────────────────────────────


@click.group()
@click.help_option("-h", "--help")
@click.version_option(__version__, "-v", "--version", prog_name="bid")
def bid() -> None:
    "Utilities for bidding."
    pass


# Cast to Group to satisfy type checker (click.group decorator returns Group)
bid_group: click.Group = bid  # type: ignore[assignment]
bid_group.add_command(init)
bid_group.add_command(setup)
bid_group.add_command(clean)
bid_group.add_command(beautify)
bid_group.add_command(combine_pdf)
# Hidden alias for combine-pdf
cpdf_alias = click.Command(
    name="cpdf",
    callback=combine_pdf.callback,
    params=combine_pdf.params,
    hidden=True,
    help=combine_pdf.help,
)
bid_group.add_command(cpdf_alias)
bid_group.add_command(word2pdf)
bid_group.add_command(audit)
bid_group.add_command(vo)
bid_group.add_command(ho)
bid_group.add_command(co)
bid_group.add_command(rate_cmd)
bid_group.add_command(mob_cmd)
bid_group.add_command(mob_config_cmd)

if __name__ == "__main__":
    bid()

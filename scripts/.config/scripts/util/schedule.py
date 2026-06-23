"""Gantt chart generation from markdown schedule files."""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, timedelta
from pathlib import Path


@dataclass
class Item:
    name: str
    lead_min: int
    lead_max: int
    origin: str | None = None
    is_milestone: bool = False
    freight_weeks_min: int = 0  # conservative low end (for display)
    freight_weeks: int = 0      # conservative high end; used for scheduling
    depends_on: list[str] = field(default_factory=list)
    start_week: int = 0  # computed by _compute_schedule
    sync: bool = False   # [sync] — back-schedule to arrive with latest item in phase


def _calc_freight(origin: str | None) -> tuple[int, int]:
    """Return (min, max) sea freight weeks based on origin country code.

    SG → 0, CN → 2 weeks (short sea), all other non-SG → 4–6 weeks (long haul).
    Returns (min, max); max is used for conservative scheduling.
    """
    if not origin:
        return (0, 0)
    country = origin.upper().split()[-1]
    if country == "SG":
        return (0, 0)
    if country == "CN":
        return (2, 2)
    return (4, 6)  # long-haul sea freight


@dataclass
class Phase:
    name: str
    items: list[Item] = field(default_factory=list)
    depends_on: list[str] = field(default_factory=list)
    fixed_start: int | None = None  # [start: WK15] → 15; overrides / floors dependency calc
    start_week: int = 0
    end_week: int = 0


@dataclass
class Schedule:
    project_name: str
    start_date: date
    phases: list[Phase] = field(default_factory=list)


_LEAD_RANGE_RE  = re.compile(r":\s*(\d+)\s*[–\-]\s*(\d+)\s*wks?(.*)$", re.IGNORECASE)
_LEAD_SINGLE_RE = re.compile(r":\s*(\d+)\s*wks?(.*)$", re.IGNORECASE)
_AFTER_RE       = re.compile(r"\[after:\s*([^\]]+)\]", re.IGNORECASE)
_PHASE_START_RE = re.compile(r"\[start:\s*(?:WK)?(\d+)\]", re.IGNORECASE)
_START_RE       = re.compile(r"^start:\s*(\d{4}-\d{2}-\d{2})", re.IGNORECASE)
_AIR_RE         = re.compile(r"\[air\]", re.IGNORECASE)
_SYNC_RE        = re.compile(r"\[sync\]", re.IGNORECASE)


def parse_schedule(md_text: str) -> Schedule:
    phases: list[Phase] = []
    project_name = ""
    start_date = date.today()
    current_phase: Phase | None = None

    for line in md_text.splitlines():
        s = line.strip()

        if s.startswith("# ") and not s.startswith("## "):
            project_name = s[2:].strip()
            continue

        m = _START_RE.match(s)
        if m:
            start_date = date.fromisoformat(m.group(1))
            continue

        if s.startswith("## "):
            content = s[3:].strip()
            depends: list[str] = []
            fixed_start: int | None = None
            # Parse [start: WKN] first so [after:] doesn't consume it
            ps_m = _PHASE_START_RE.search(content)
            if ps_m:
                fixed_start = int(ps_m.group(1)) - 1  # convert 1-based WK to 0-based
                content = (content[: ps_m.start()] + content[ps_m.end():]).strip()
            after_m = _AFTER_RE.search(content)
            if after_m:
                depends = [d.strip() for d in after_m.group(1).split(",")]
                content = content[: after_m.start()].strip()
            current_phase = Phase(name=content, depends_on=depends, fixed_start=fixed_start)
            phases.append(current_phase)
            continue

        if s.startswith("- ") and current_phase is not None:
            text = s[2:].strip()

            if text.lower().startswith("milestone:"):
                milestone_name = text[len("milestone:"):].strip()
                current_phase.items.append(
                    Item(name=milestone_name, lead_min=0, lead_max=0, is_milestone=True)
                )
                continue

            # Strip [after:] and [air] before origin capture so they aren't treated as origin text
            item_deps: list[str] = []
            after_m = _AFTER_RE.search(text)
            if after_m:
                item_deps = [d.strip() for d in after_m.group(1).split(",")]
                text = (text[: after_m.start()] + text[after_m.end():]).strip()
            air_m = _AIR_RE.search(text)
            air = bool(air_m)
            if air_m:
                text = (text[: air_m.start()] + text[air_m.end():]).strip()
            sync_m = _SYNC_RE.search(text)
            sync = bool(sync_m)
            if sync_m:
                text = (text[: sync_m.start()] + text[sync_m.end():]).strip()

            m2 = _LEAD_RANGE_RE.search(text)
            if m2:
                lead_min, lead_max = int(m2.group(1)), int(m2.group(2))
                origin = m2.group(3).strip() or None
                name = text[: m2.start()].strip()
            else:
                m3 = _LEAD_SINGLE_RE.search(text)
                if m3:
                    lead_min = lead_max = int(m3.group(1))
                    origin = m3.group(2).strip() or None
                    name = text[: m3.start()].strip()
                else:
                    name, lead_min, lead_max, origin = text, 0, 0, None

            if air:
                fmin, fmax = 2, 2
            else:
                fmin, fmax = _calc_freight(origin)
            current_phase.items.append(
                Item(name=name, lead_min=lead_min, lead_max=lead_max,
                     origin=origin, freight_weeks_min=fmin, freight_weeks=fmax,
                     depends_on=item_deps, sync=sync)
            )

    return Schedule(project_name=project_name, start_date=start_date, phases=phases)


def _compute_schedule(phases: list[Phase]) -> None:
    by_phase: dict[str, Phase] = {p.name: p for p in phases}
    # Item lookup: item name → item; phase names take priority if name collision
    by_item: dict[str, Item] = {
        item.name: item
        for p in phases
        for item in p.items
        if not item.is_milestone
    }

    def _item_delivery(item: Item) -> int:
        return item.start_week + item.lead_max + item.freight_weeks

    def _dep_end(name: str) -> int:
        if name in by_phase:
            return by_phase[name].end_week
        if name in by_item:
            return _item_delivery(by_item[name])
        return 0  # unknown reference, ignored

    for i, phase in enumerate(phases):
        if phase.fixed_start is not None and not phase.depends_on:
            # [start: WKN] with no [after:] → hard override; ignore implicit prev-phase sequencing
            phase.start_week = phase.fixed_start
        elif phase.depends_on:
            dep_end = max((_dep_end(d) for d in phase.depends_on), default=0)
            # [after: X] [start: WKN] → later of dependency end or fixed floor
            phase.start_week = max(dep_end, phase.fixed_start) if phase.fixed_start is not None else dep_end
        else:
            # No annotation → sequential after previous phase
            phase.start_week = 0 if i == 0 else phases[i - 1].end_week

        # Pass 1: compute start weeks for non-sync items
        for item in phase.items:
            if item.is_milestone or item.sync:
                continue
            if item.depends_on:
                item.start_week = max((_dep_end(d) for d in item.depends_on), default=phase.start_week)
            else:
                item.start_week = phase.start_week

        # Pass 2: back-schedule sync items to arrive with the latest non-sync delivery
        non_sync = [it for it in phase.items if not it.is_milestone and not it.sync]
        max_delivery = max((_item_delivery(it) for it in non_sync), default=phase.start_week)
        for item in phase.items:
            if item.is_milestone or not item.sync:
                continue
            item.start_week = max(phase.start_week, max_delivery - item.lead_max - item.freight_weeks)

        phase.end_week = max(
            (_item_delivery(item) for item in phase.items if not item.is_milestone),
            default=phase.start_week,
        )


def generate_excel(schedule: Schedule, output_path: Path) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _compute_schedule(schedule.phases)

    total_weeks = max((p.end_week for p in schedule.phases), default=12) + 2
    total_weeks = max(total_weeks, 12)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"

    # ── Styles ────────────────────────────────────────────────────────────────
    # Aptos is the default Microsoft 365 font (2023+); Excel substitutes Arial on older installs.
    def F(**kw) -> Font:
        return Font(name="Aptos", **kw)

    # Palette derived from Jason Blue 005BBF
    # Dark shade (~70%): 003D80  — phase header background, title
    # Dominant:           005BBF  — column header, phase span bar
    # Medium tint (~35%): 5594D4  — odd item bars
    # Light tint (~60%):  99BDE5  — even item bars
    # Very light (~85%):  D9E6F5  — alt row background
    PHASE_FONT  = F(bold=True, color="FFFFFF", size=11)
    PHASE_FILL  = PatternFill("solid", fgColor="003D80")
    PHASE_BAR   = PatternFill("solid", fgColor="005BBF")
    HDR_FONT    = F(bold=True, size=9, color="FFFFFF")
    HDR_FILL    = PatternFill("solid", fgColor="005BBF")
    ITEM_FONT   = F(size=11)
    MS_FONT     = F(bold=True, size=11, color="833C00")
    MS_BAR      = PatternFill("solid", fgColor="F4B942")
    # Two alternating bar shades so adjacent items are visually distinct
    BAR_FILLS     = [
        PatternFill("solid", fgColor="99BDE5"),  # even rows: light tint
        PatternFill("solid", fgColor="5594D4"),  # odd rows:  medium tint
    ]
    # Sea freight band: soft green — reads as "in transit / logistics"
    FREIGHT_FILLS = [
        PatternFill("solid", fgColor="A9D18E"),  # even rows: light green
        PatternFill("solid", fgColor="70AD47"),  # odd rows:  medium green
    ]
    ALT_FILL      = PatternFill("solid", fgColor="D9E6F5")
    CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    INDENT1     = Alignment(horizontal="left", vertical="center", indent=1)
    INDENT2     = Alignment(horizontal="left", vertical="center", indent=2)

    FIXED = 5   # Name | Lead Time | Origin | Start Wk | End Wk
    WK1   = FIXED + 1

    # ── Row 1: project title ─────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=FIXED + total_weeks)
    tc = ws.cell(row=1, column=1, value=f"Project Schedule: {schedule.project_name}")
    tc.font = F(bold=True, size=14, color="003D80")
    tc.alignment = INDENT1
    ws.row_dimensions[1].height = 24

    # ── Row 2: column headers ─────────────────────────────────────────────────
    for ci, h in enumerate(["Item", "Lead Time", "Origin", "Start\nWk", "End\nWk"], 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER

    for w in range(1, total_weeks + 1):
        wdate = schedule.start_date + timedelta(weeks=w - 1)
        c = ws.cell(row=2, column=WK1 + w - 1, value=f"W{w}\n{wdate.strftime('%d %b')}")
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
    ws.row_dimensions[2].height = 40  # taller to show two-line date clearly

    # ── Data rows ─────────────────────────────────────────────────────────────
    row = 3
    for phase in schedule.phases:
        # Phase header
        for ci in range(1, FIXED + total_weeks + 1):
            ws.cell(row=row, column=ci).fill = PHASE_FILL
        ws.cell(row=row, column=1, value=phase.name).font = PHASE_FONT
        ws.cell(row=row, column=1).alignment = INDENT1
        if phase.start_week < phase.end_week:
            for ci in range(WK1 + phase.start_week, min(WK1 + phase.end_week, WK1 + total_weeks)):
                ws.cell(row=row, column=ci).fill = PHASE_BAR
        ws.row_dimensions[row].height = 18
        row += 1

        for j, item in enumerate(phase.items):
            bg = ALT_FILL if j % 2 == 0 else None
            bar_fill = BAR_FILLS[j % 2]

            if item.is_milestone:
                c = ws.cell(row=row, column=1, value=f"◆  {item.name}")
                c.font = MS_FONT
                c.alignment = INDENT2
                # Position: start of phase if no non-milestone items precede this one, else end
                at_start = not any(not it.is_milestone for it in phase.items[:j])
                if at_start:
                    ms_week = phase.start_week
                else:
                    ms_week = (phase.end_week - 1) if phase.end_week > phase.start_week else phase.start_week
                ms_col = min(WK1 + ms_week, WK1 + total_weeks - 1)
                mc = ws.cell(row=row, column=ms_col)
                mc.value = "◆"
                mc.fill = MS_BAR
                mc.alignment = CENTER
                mc.font = F(bold=True, color="833C00")
            else:
                c = ws.cell(row=row, column=1, value=item.name)
                c.font = ITEM_FONT
                c.alignment = INDENT2

                lt = (f"{item.lead_max} wks" if item.lead_min == item.lead_max
                      else f"{item.lead_min}–{item.lead_max} wks") if item.lead_max else "–"
                delivery_max = item.start_week + item.lead_max + item.freight_weeks
                delivery_min = item.start_week + item.lead_max + item.freight_weeks_min
                delivery = (f"WK {delivery_max}" if delivery_min == delivery_max
                            else f"WK {delivery_min}–{delivery_max}")
                for ci, val in [
                    (2, lt),
                    (3, item.origin or ""),
                    (4, item.start_week + 1),
                    (5, delivery),
                ]:
                    cell = ws.cell(row=row, column=ci, value=val)
                    cell.font = ITEM_FONT
                    cell.alignment = CENTER

                if item.lead_max > 0:
                    for ci in range(
                        WK1 + item.start_week,
                        min(WK1 + item.start_week + item.lead_max, WK1 + total_weeks),
                    ):
                        ws.cell(row=row, column=ci).fill = bar_fill

                if item.freight_weeks > 0:
                    freight_start = WK1 + item.start_week + item.lead_max
                    for ci in range(
                        freight_start,
                        min(freight_start + item.freight_weeks, WK1 + total_weeks),
                    ):
                        ws.cell(row=row, column=ci).fill = FREIGHT_FILLS[j % 2]

            if bg:
                for ci in range(1, FIXED + 1):
                    cell = ws.cell(row=row, column=ci)
                    if cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                        cell.fill = bg

            ws.row_dimensions[row].height = 16
            row += 1

        # spacer
        ws.row_dimensions[row].height = 5
        row += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 7
    ws.column_dimensions["E"].width = 7
    for w in range(1, total_weeks + 1):
        ws.column_dimensions[get_column_letter(WK1 + w - 1)].width = 8.5  # wide enough for "01 Aug"

    ws.freeze_panes = ws.cell(row=3, column=WK1)

    wb.save(output_path)
